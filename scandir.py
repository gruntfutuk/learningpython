from openpyxl import Workbook
import os
from datetime import datetime

workbook = Workbook()
sheet = workbook.create_sheet("Totals")
# sheet = workbook.add_sheet("Totals",cell_overwrite_ok=True)



envs = ['.']

for env in envs:
    files = [(env, pathname, datetime.time.strftime('%Y/%m/%d', datetime.time.gmtime(os.path.getctime(pathname))), os.path.getsize(pathname)) for pathname in os.listdir(env) if os.path.isfile(pathname)]

row = 1
col = -2
envprev = ""
for file in files:
    print(f'writing {files}')  
    if file[0] != envprev:
        row = 1
        col += 3
        count = 0
        envprev = file[0]
        print(f'row: {row}, col: {col}, value: {file[0]}')
        sheet.cell(row=row, column=col).value = file[0]
    else:
        row += 1
        print(f'row: {row}, col: {col}, values: {file[1]}, {file[2]}')
        sheet.cell(row=row, column=col).value = file[1]
        sheet.cell(row=row, column=col+1).value = file[2]

workbook.save("totals.xls")